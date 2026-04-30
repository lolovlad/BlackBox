from __future__ import annotations

import base64
import binascii
import csv
import json
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Blueprint, abort, current_app, jsonify, render_template, request, send_file, url_for
from flask_login import login_required
from sqlalchemy import inspect

from src.database import Alarms, Samples, Video, db
from src.webui.data_labels import (
    all_analog_keys,
    all_discrete_keys,
    analog_labels_for,
    discrete_labels_for,
)
from src.webui.modbus_service import analog_discrete_for_csv, decode_to_processed
from src.webui.paths import TEMPLATES_DIR
from src.webui.repositories.data_repository import DataRepository
from src.webui.services.data_service import TABLE_PAGE_SIZE, DataService, parse_data_filter
from src.webui.system_settings import read_env_file
from src.webui.timezone_utils import (
    configured_timezone_name,
    format_in_configured_timezone,
    now_in_configured_timezone_naive,
    to_configured_timezone,
)

data_router = Blueprint("data", __name__, url_prefix="/data", template_folder=str(TEMPLATES_DIR))
DATETIME_UI_FORMAT = "%d.%m.%Y %H:%M:%S"
DATETIME_API_FORMAT = "%Y-%m-%dT%H:%M:%S"


def _service() -> DataService:
    session_factory = current_app.extensions["session_factory"]
    collector = current_app.extensions["modbus_collector"]
    repo = DataRepository(session_factory)
    try:
        gpio_enabled = bool(inspect(db.engine).has_table("alarms_raspberry"))
    except Exception:
        gpio_enabled = True
    return DataService(repo, alarms_enabled=collector._alarms_enabled, gpio_alarms_enabled=gpio_enabled)


@data_router.route("/", methods=["GET"])
@login_required
def page():
    analog_opts = analog_labels_for(all_analog_keys())
    discrete_opts = discrete_labels_for(all_discrete_keys())
    return render_template(
        "data/index.html",
        analog_options=analog_opts,
        discrete_options=discrete_opts,
        table_page_size=TABLE_PAGE_SIZE,
    )


@data_router.route("/tables", methods=["GET"])
@login_required
def tables():
    flt = parse_data_filter(request.args)
    ctx = _service().collect_tab(flt)
    return render_template("data/_tables.html", **ctx)


def _video_allowed_roots() -> list[Path]:
    env = read_env_file(current_app.extensions["env_path"])
    raw = (env.get("VIDEO_STORAGE_DIR") or "").strip()
    if not raw:
        return []
    root = Path(raw).expanduser().resolve()
    roots = [root]
    if re.fullmatch(r"cam\d+", root.name.lower()) and root.parent.is_dir():
        roots.append(root.parent.resolve())
    uniq: list[Path] = []
    for p in roots:
        if p not in uniq:
            uniq.append(p)
    return uniq


def _is_under_any_root(path: Path, roots: list[Path]) -> bool:
    if not roots:
        return True
    for root in roots:
        try:
            path.relative_to(root)
            return True
        except ValueError:
            continue
    return False


def _normalize_video_fs_path(raw: str) -> Path:
    value = str(raw or "").strip()
    if value.lower().startswith("file="):
        value = value.split("=", 1)[1].strip()
    if value.lower().startswith("file://"):
        value = value[7:].strip()
    return Path(value).expanduser()


@data_router.route("/videos/<int:video_id>/download", methods=["GET"])
@login_required
def download_video(video_id: int):
    session_factory = current_app.extensions["session_factory"]
    with session_factory() as session:
        item = session.get(Video, video_id)
    if item is None:
        abort(404)

    file_path = _normalize_video_fs_path(item.file_path)
    try:
        resolved = file_path.resolve()
    except OSError:
        abort(404)
    if not resolved.exists() or not resolved.is_file():
        abort(404)
    if not _is_under_any_root(resolved, _video_allowed_roots()):
        abort(403)

    return send_file(resolved, as_attachment=True, download_name=(item.file_name or resolved.name))


@data_router.route("/videos/<int:video_id>/open", methods=["GET"])
@login_required
def open_video(video_id: int):
    session_factory = current_app.extensions["session_factory"]
    with session_factory() as session:
        item = session.get(Video, video_id)
    if item is None:
        abort(404)

    file_path = _normalize_video_fs_path(item.file_path)
    try:
        resolved = file_path.resolve()
    except OSError:
        abort(404)

    if not resolved.exists() or not resolved.is_file():
        abort(404)
    if not _is_under_any_root(resolved, _video_allowed_roots()):
        abort(403)

    # Отдаём видео как inline-контент, чтобы его можно было открыть прямо в браузере.
    return send_file(resolved, as_attachment=False, download_name=(item.file_name or resolved.name))


@data_router.route("/export", methods=["POST"])
@login_required
def export_submit():
    flt = parse_data_filter(request.form)
    if flt.date_from is None or flt.date_to is None:
        return render_template(
            "data/_export_result.html",
            error="Для экспорта укажите дату/время начала и конца.",
            download_url=None,
        )
    static_csv_dir: Path = current_app.extensions["static_csv_dir"]
    svc = _service()
    try:
        export_path = svc.build_export(flt, static_csv_dir)
    except ValueError as exc:
        if str(exc) == "alarms_disabled":
            return render_template(
                "data/_export_result.html",
                error="Таблица аварий недоступна (нет таблицы в БД). Выполните миграции.",
                download_url=None,
            )
        raise
    rel = export_path.relative_to(current_app.static_folder)
    download_url = url_for("static", filename=str(rel).replace("\\", "/"))
    return render_template("data/_export_result.html", error=None, download_url=download_url)


def _selected_export_tables() -> list[str]:
    out: list[str] = []
    if request.form.get("table_analog") == "1":
        out.append("analog")
    if request.form.get("table_discrete") == "1":
        out.append("discrete")
    if request.form.get("table_alarms") == "1":
        out.append("alarms")
    return out


def _normalize_export_range(date_from: datetime | None, date_to: datetime | None) -> tuple[datetime | None, datetime | None]:
    if date_from is not None and date_to is not None and date_from > date_to:
        return date_to, date_from
    return date_from, date_to


@data_router.route("/export/batch", methods=["POST"])
@login_required
def export_batch():
    tables = _selected_export_tables()
    if not tables:
        return render_template("data/_export_result.html", error="Выберите хотя бы одну таблицу для экспорта.", download_url=None)

    export_format = (request.form.get("export_format") or "csv_zip").strip().lower()
    if export_format not in ("csv_zip", "excel"):
        return render_template("data/_export_result.html", error="Некорректный формат экспорта.", download_url=None)

    date_from, date_to = _normalize_export_range(
        _parse_dt_local(request.form.get("date_from")),
        _parse_dt_local(request.form.get("date_to")),
    )
    if date_from is None or date_to is None:
        return render_template("data/_export_result.html", error="Для экспорта укажите дату/время начала и конца.", download_url=None)
    sort_desc = (request.form.get("sort") or "desc").lower() != "asc"
    analog_keys = request.form.getlist("analog_col_export")
    discrete_keys = request.form.getlist("discrete_col_export")
    if "analog" in tables and not analog_keys:
        return render_template("data/_export_result.html", error="Для таблицы «Аналоги» выберите хотя бы одно поле.", download_url=None)
    if "discrete" in tables and not discrete_keys:
        return render_template("data/_export_result.html", error="Для таблицы «Дискреты» выберите хотя бы одно поле.", download_url=None)

    collector = current_app.extensions["modbus_collector"]
    if "alarms" in tables and not collector._alarms_enabled:
        return render_template(
            "data/_export_result.html",
            error="Таблица аварий недоступна (нет таблицы в БД). Выполните миграции.",
            download_url=None,
        )

    static_csv_dir: Path = current_app.extensions["static_csv_dir"]
    static_csv_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    session_factory = current_app.extensions["session_factory"]

    if export_format == "csv_zip":
        out_path = static_csv_dir / f"export_package_{stamp}.zip"
        try:
            with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                with session_factory() as session:
                    # Если экспортируем аварии — добавляем в пакет и связанные видео.
                    allowed_video_roots = _video_allowed_roots() if "alarms" in tables else []
                    videos_csv_tmp_path: Path | None = None
                    if "alarms" in tables:
                        alarm_ids = [
                            r[0]
                            for r in session.query(Alarms.id)
                            .filter(Alarms.created_at >= date_from, Alarms.created_at <= date_to)
                            .all()
                        ]
                        video_q = (
                            session.query(Video)
                            .filter(Video.alarm_id.is_not(None))
                            .filter(Video.alarm_id.in_(alarm_ids))
                            .order_by(Video.created_at.desc() if sort_desc else Video.created_at.asc())
                        )
                        videos_csv_tmp = tempfile.NamedTemporaryFile(
                            prefix="videos_", suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8-sig"
                        )
                        videos_csv_tmp_path = Path(videos_csv_tmp.name)
                        try:
                            vw = csv.writer(videos_csv_tmp, delimiter=";")
                            vw.writerow(
                                ["ID", "Дата", "Время", "Время_захвата", "Имя_файла", "Путь_к_файлу", "ID_аварии"]
                            )
                            for v in video_q.yield_per(1000):
                                vw.writerow(
                                    [
                                        v.id,
                                        v.created_at.strftime("%d/%m/%Y"),
                                        v.created_at.strftime("%H:%M:%S"),
                                        v.captured_at.strftime("%d/%m/%Y %H:%M:%S"),
                                        v.file_name,
                                        v.file_path,
                                        v.alarm_id if v.alarm_id is not None else "",
                                    ]
                                )

                                # Подключаем бинарные файлы в ZIP.
                                file_path = _normalize_video_fs_path(v.file_path)
                                try:
                                    resolved = file_path.resolve()
                                except OSError:
                                    continue
                                if not resolved.exists() or not resolved.is_file():
                                    continue
                                if not _is_under_any_root(resolved, allowed_video_roots):
                                    continue
                                arcname = f"videos/{v.id}_{v.file_name}"
                                zf.write(str(resolved), arcname=arcname)
                        finally:
                            videos_csv_tmp.close()

                    for table in tables:
                        csv_tmp = tempfile.NamedTemporaryFile(
                            prefix=f"{table}_", suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8-sig"
                        )
                        csv_tmp_path = Path(csv_tmp.name)
                        try:
                            writer = csv.writer(csv_tmp, delimiter=";")
                            if table == "analog":
                                writer.writerow(["Дата", "Время", *analog_keys])
                                stmt = session.query(Samples).filter(
                                    Samples.created_at >= date_from, Samples.created_at <= date_to
                                )
                                stmt = stmt.order_by(Samples.created_at.desc() if sort_desc else Samples.created_at.asc())
                                for item in stmt.yield_per(1000):
                                    processed = decode_to_processed(item.date)
                                    analog, _ = analog_discrete_for_csv(processed)
                                    writer.writerow(
                                        [
                                            item.created_at.strftime("%d/%m/%Y"),
                                            item.created_at.strftime("%H:%M:%S"),
                                            *[analog.get(k, "") for k in analog_keys],
                                        ]
                                    )
                            elif table == "discrete":
                                writer.writerow(["Дата", "Время", *discrete_keys])
                                stmt = session.query(Samples).filter(
                                    Samples.created_at >= date_from, Samples.created_at <= date_to
                                )
                                stmt = stmt.order_by(Samples.created_at.desc() if sort_desc else Samples.created_at.asc())
                                for item in stmt.yield_per(1000):
                                    processed = decode_to_processed(item.date)
                                    _, discrete = analog_discrete_for_csv(processed)
                                    writer.writerow(
                                        [
                                            item.created_at.strftime("%d/%m/%Y"),
                                            item.created_at.strftime("%H:%M:%S"),
                                            *[1 if bool(discrete.get(k, False)) else 0 for k in discrete_keys],
                                        ]
                                    )
                            else:
                                writer.writerow(["Дата", "Время", "Название", "Состояние"])
                                stmt = session.query(Alarms).filter(
                                    Alarms.created_at >= date_from, Alarms.created_at <= date_to
                                )
                                stmt = stmt.order_by(Alarms.created_at.desc() if sort_desc else Alarms.created_at.asc())
                                for item in stmt.yield_per(1000):
                                    writer.writerow(
                                        [
                                            item.created_at.strftime("%d/%m/%Y"),
                                            item.created_at.strftime("%H:%M:%S"),
                                            item.name,
                                            getattr(item, "state", "active"),
                                        ]
                                    )
                        finally:
                            csv_tmp.close()
                        zf.write(csv_tmp_path, arcname=f"{table}.csv")
                        csv_tmp_path.unlink(missing_ok=True)
                        if videos_csv_tmp_path is not None:
                            zf.write(str(videos_csv_tmp_path), arcname="videos.csv")
                            videos_csv_tmp_path.unlink(missing_ok=True)
        except Exception:
            out_path.unlink(missing_ok=True)
            raise
        rel = out_path.relative_to(current_app.static_folder)
        return render_template(
            "data/_export_result.html",
            error=None,
            download_url=url_for("static", filename=str(rel).replace("\\", "/")),
        )

    try:
        from openpyxl import Workbook
    except Exception:
        return render_template(
            "data/_export_result.html",
            error="Для Excel-экспорта требуется пакет openpyxl. Установите зависимость и повторите.",
            download_url=None,
        )

    out_path = static_csv_dir / f"export_package_{stamp}.xlsx"
    try:
        wb = Workbook()
        wb.remove(wb.active)
        sheet_title = {"analog": "Аналоги", "discrete": "Дискреты", "alarms": "Аварии", "videos": "Видео"}
        with session_factory() as session:
            for table in tables:
                ws = wb.create_sheet(sheet_title.get(table, table))
                if table == "analog":
                    ws.append(["Дата", "Время", *analog_keys])
                    stmt = session.query(Samples).filter(Samples.created_at >= date_from, Samples.created_at <= date_to)
                    stmt = stmt.order_by(Samples.created_at.desc() if sort_desc else Samples.created_at.asc())
                    for item in stmt.yield_per(1000):
                        processed = decode_to_processed(item.date)
                        analog, _ = analog_discrete_for_csv(processed)
                        ws.append(
                            [
                                item.created_at.strftime("%d/%m/%Y"),
                                item.created_at.strftime("%H:%M:%S"),
                                *[analog.get(k, "") for k in analog_keys],
                            ]
                        )
                elif table == "discrete":
                    ws.append(["Дата", "Время", *discrete_keys])
                    stmt = session.query(Samples).filter(Samples.created_at >= date_from, Samples.created_at <= date_to)
                    stmt = stmt.order_by(Samples.created_at.desc() if sort_desc else Samples.created_at.asc())
                    for item in stmt.yield_per(1000):
                        processed = decode_to_processed(item.date)
                        _, discrete = analog_discrete_for_csv(processed)
                        ws.append(
                            [
                                item.created_at.strftime("%d/%m/%Y"),
                                item.created_at.strftime("%H:%M:%S"),
                                *[1 if bool(discrete.get(k, False)) else 0 for k in discrete_keys],
                            ]
                        )
                else:
                    ws.append(["Дата", "Время", "Название", "Состояние"])
                    stmt = session.query(Alarms).filter(Alarms.created_at >= date_from, Alarms.created_at <= date_to)
                    stmt = stmt.order_by(Alarms.created_at.desc() if sort_desc else Alarms.created_at.asc())
                    for item in stmt.yield_per(1000):
                        ws.append(
                            [
                                item.created_at.strftime("%d/%m/%Y"),
                                item.created_at.strftime("%H:%M:%S"),
                                item.name,
                                getattr(item, "state", "active"),
                            ]
                        )

            if "alarms" in tables:
                ws = wb.create_sheet(sheet_title["videos"])
                ws.append(["ID", "Сохранено_в_БД", "Время_захвата", "Имя_файла", "ID_аварии", "Путь_к_файлу"])
                video_stmt = (
                    session.query(Video)
                    .join(Alarms, Video.alarm_id == Alarms.id)
                    .filter(Alarms.created_at >= date_from, Alarms.created_at <= date_to)
                    .order_by(Video.created_at.desc() if sort_desc else Video.created_at.asc())
                )
                for v in video_stmt.yield_per(1000):
                    ws.append(
                        [
                            v.id,
                            f"{v.created_at.strftime('%d/%m/%Y')} {v.created_at.strftime('%H:%M:%S')}",
                            f"{v.captured_at.strftime('%d/%m/%Y %H:%M:%S')}",
                            v.file_name,
                            v.alarm_id,
                            v.file_path,
                        ]
                    )
        wb.save(out_path)
    except Exception:
        out_path.unlink(missing_ok=True)
        raise
    rel = out_path.relative_to(current_app.static_folder)
    return render_template(
        "data/_export_result.html",
        error=None,
        download_url=url_for("static", filename=str(rel).replace("\\", "/")),
    )


def _parse_dt_local(raw: str | None) -> datetime | None:
    if raw is None:
        return None
    val = str(raw).strip()
    if not val:
        return None
    try:
        return datetime.fromisoformat(val)
    except ValueError:
        return None


@data_router.route("/charts", methods=["GET"])
@login_required
def charts_page():
    analog_opts = analog_labels_for(all_analog_keys())
    discrete_opts = discrete_labels_for(all_discrete_keys())
    return render_template(
        "data/charts.html",
        analog_options=analog_opts,
        discrete_options=discrete_opts,
        chart_app_timezone=configured_timezone_name(),
    )


def _requested_chart_columns(table: str) -> list[str]:
    requested: list[str]
    raw_b64 = request.args.get("selected_col_b64")
    if raw_b64:
        requested = _decode_selected_columns(raw_b64)
    else:
        requested = request.args.getlist("analog_col") if table == "analog" else request.args.getlist("discrete_col")
    if table == "analog":
        allowed = set(all_analog_keys())
    else:
        allowed = set(all_discrete_keys())
    return [col for col in requested if col in allowed]


def _decode_selected_columns(raw: str | None) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    padded = text + ("=" * ((4 - len(text) % 4) % 4))
    try:
        decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        payload = json.loads(decoded)
        if not isinstance(payload, list):
            return []
        return [str(v) for v in payload]
    except (ValueError, UnicodeDecodeError, binascii.Error):
        return []


def _collect_second_points(rows: list[Any], table: str, columns: list[str]) -> list[dict[str, Any]]:
    by_second: dict[datetime, dict[str, Any]] = {}
    for item in rows:
        sec = item.created_at.replace(microsecond=0)
        processed = decode_to_processed(item.date)
        analog, discrete = analog_discrete_for_csv(processed)
        values: dict[str, Any] = {}
        for col in columns:
            if table == "analog":
                values[col] = analog.get(col, None)
            else:
                values[col] = 1 if bool(discrete.get(col, False)) else 0
        # Если в одну секунду несколько замеров, берем последний
        by_second[sec] = values

    points: list[dict[str, Any]] = []
    for sec in sorted(by_second.keys()):
        date_label = format_in_configured_timezone(sec, "%d.%m.%Y")
        time_label = format_in_configured_timezone(sec, "%H:%M")
        ts_ms = int(to_configured_timezone(sec).timestamp() * 1000)
        points.append(
            {
                "ts": sec.strftime(DATETIME_API_FORMAT),
                "ts_ms": ts_ms,
                "date_label": date_label,
                "time_label": time_label,
                "label": f"{date_label}\n{time_label}",
                "values": by_second[sec],
            }
        )
    return points


def _parse_since(raw: str | None) -> datetime | None:
    if raw is None:
        return None
    val = str(raw).strip()
    if not val:
        return None
    for fmt in (DATETIME_API_FORMAT, DATETIME_UI_FORMAT):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(val)
    except ValueError:
        return None


@data_router.route("/charts/api/init", methods=["GET"])
@login_required
def charts_api_init():
    table = (request.args.get("table") or "analog").strip().lower()
    if table not in ("analog", "discrete"):
        table = "analog"

    date_from = _parse_dt_local(request.args.get("date_from"))
    date_to = _parse_dt_local(request.args.get("date_to"))
    if date_from is not None and date_to is not None and date_from > date_to:
        date_from, date_to = date_to, date_from

    realtime = date_from is None and date_to is None
    if realtime:
        now = now_in_configured_timezone_naive()
        date_from = now.replace(hour=0, minute=0, second=0, microsecond=0)
        date_to = now

    repo = DataRepository(current_app.extensions["session_factory"])
    rows = (
        repo.list_analogs(created_from=date_from, created_to=date_to, sort_desc=False, limit=None)
        if table == "analog"
        else repo.list_discretes(created_from=date_from, created_to=date_to, sort_desc=False, limit=None)
    )

    columns = _requested_chart_columns(table)
    points = _collect_second_points(rows, table, columns)
    labels = dict(analog_labels_for(columns) if table == "analog" else discrete_labels_for(columns))
    return jsonify(
        {
            "table": table,
            "columns": columns,
            "column_labels": labels,
            "points": points,
            "last_ts": points[-1]["ts"] if points else None,
            "row_count": len(points),
            "realtime": realtime,
        }
    )


@data_router.route("/charts/api/update", methods=["GET"])
@login_required
def charts_api_update():
    table = (request.args.get("table") or "analog").strip().lower()
    if table not in ("analog", "discrete"):
        table = "analog"
    columns = _requested_chart_columns(table)
    since = _parse_since(request.args.get("since"))

    now = now_in_configured_timezone_naive()
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    created_from = since if since is not None else day_start
    if created_from < day_start:
        created_from = day_start

    repo = DataRepository(current_app.extensions["session_factory"])
    rows = (
        repo.list_analogs(created_from=created_from, created_to=now, sort_desc=False, limit=None)
        if table == "analog"
        else repo.list_discretes(created_from=created_from, created_to=now, sort_desc=False, limit=None)
    )
    points = _collect_second_points(rows, table, columns)

    if since is not None:
        filtered: list[dict[str, Any]] = []
        for point in points:
            pt = _parse_since(point["ts"])
            if pt is not None and pt > since:
                filtered.append(point)
        points = filtered

    labels = dict(analog_labels_for(columns) if table == "analog" else discrete_labels_for(columns))
    return jsonify(
        {
            "table": table,
            "columns": columns,
            "column_labels": labels,
            "points": points,
            "last_ts": points[-1]["ts"] if points else (since.strftime(DATETIME_API_FORMAT) if since else None),
            "row_count": len(points),
            "realtime": True,
        }
    )
