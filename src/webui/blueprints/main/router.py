from __future__ import annotations

import base64
import binascii
import csv
import io
import json
import re
import os
import shutil
import time
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from flask import Blueprint, Response, current_app, flash, redirect, render_template, render_template_string, request, url_for
from flask_login import current_user, login_required
from jinja2 import TemplateNotFound

from src.webui.auth_utils import admin_required
from src.webui.data_labels import (
    all_analog_keys,
    all_discrete_keys,
    analog_labels_for,
    discrete_labels_for,
    filter_valid_analog,
    filter_valid_discrete,
)
from src.webui.paths import TEMPLATES_DIR
from src.webui.emergency_rule_validation import validate_emergency_rule_expression
from src.webui.repositories.data_repository import DataRepository
from src.webui.repositories.emergency_repository import EmergencyRepository
from src.database import Alarms, Video
from src.webui.modbus_service import analog_discrete_for_csv, configure_settings_path, decode_to_processed
from src.webui.modbus_service import reload_settings_cache

from src.webui.app_runtime_config import (
    APP_RUNTIME_FILENAME,
    ROOT_ENV_DEFAULTS,
    apply_app_runtime_to_environ,
    build_runtime_config,
    io_form_to_runtime,
    load_app_runtime,
    save_app_runtime,
)
from src.webui.system_settings import (
    MINIMAL_PARSER_SETTINGS_JSON,
    is_valid_parser_settings_file,
    prune_parser_settings_json_files,
    read_env_file,
    repair_parser_settings_path,
    test_modbus_settings,
    validate_parser_json,
)
from src.webui.gpio_settings import ensure_gpio_inputs_file, gpio_inputs_path, validate_gpio_inputs_json
from src.webui.timezone_utils import format_in_configured_timezone

main_router = Blueprint("main", __name__, template_folder=str(TEMPLATES_DIR))
DATETIME_UI_FORMAT = "%d.%m.%Y %H:%M:%S"


def _is_admin() -> bool:
    tu = getattr(current_user, "type_user", None)
    return bool(tu is not None and getattr(tu, "system_name", "") == "admin")


def _timezone_form_fields(
    env_map: dict[str, str], timezone_choices: list[tuple[str, str]]
) -> tuple[str, str]:
    """Значение для select и для поля «своя зона» (ключ APP_TIMEZONE в env_map)."""
    current = (env_map.get("APP_TIMEZONE") or "").strip() or "Europe/Moscow"
    tz_keys = {t[0] for t in timezone_choices}
    if current in tz_keys:
        return current, ""
    first = timezone_choices[0][0] if timezone_choices else current
    return first, current


def _project_root() -> Path:
    return Path(current_app.config["PROJECT_ROOT"])


def _ensure_parser_settings_coherent() -> None:
    root = _project_root()
    sd = _settings_dir()
    sd.mkdir(parents=True, exist_ok=True)
    prune_parser_settings_json_files(sd)
    cfg = current_app.extensions.get("app_runtime_config")
    if cfg is None:
        return
    new_cfg, changed = repair_parser_settings_path(root, cfg, sd)
    if not changed:
        return
    apply_app_runtime_to_environ(new_cfg)
    current_app.extensions["app_runtime_config"] = new_cfg
    current_app.config["PARSER_SETTINGS_PATH"] = new_cfg.parser_settings_path
    abs_p = (root / Path(new_cfg.parser_settings_path)).resolve()
    configure_settings_path(abs_p)
    reload_settings_cache()


def _effective_parser_settings_path() -> Path:
    _ensure_parser_settings_coherent()
    raw = current_app.config.get("PARSER_SETTINGS_PATH", "settings/settings.json")
    p = Path(raw)
    root = _project_root()
    resolved = p.resolve() if p.is_absolute() else (root / p).resolve()
    resolved.parent.mkdir(parents=True, exist_ok=True)
    if not resolved.exists():
        resolved.write_text(MINIMAL_PARSER_SETTINGS_JSON + "\n", encoding="utf-8")
    return resolved


def _settings_dir() -> Path:
    return _project_root() / "settings"


def _rel_parser_settings_env_name(filename: str) -> str:
    return (Path("settings") / Path(filename).name).as_posix()


def _settings_file_choices(active_basename: str) -> tuple[list[str], str | None]:
    settings_dir = _settings_dir()
    settings_dir.mkdir(parents=True, exist_ok=True)
    valid: list[str] = []
    for p in sorted(settings_dir.glob("*.json"), key=lambda x: x.name.lower()):
        if p.name == APP_RUNTIME_FILENAME:
            continue
        if is_valid_parser_settings_file(p):
            valid.append(p.name)
    warning: str | None = None
    if "settings.json" in valid and valid[0] != "settings.json":
        valid.remove("settings.json")
        valid.insert(0, "settings.json")
    return valid, warning


def _set_active_parser_settings(filename: str) -> Path:
    safe_name = Path(filename).name
    sd = _settings_dir().resolve()
    target = (sd / safe_name).resolve()
    if not target.exists():
        raise FileNotFoundError(f"Файл настроек не найден: {safe_name}")
    try:
        target.relative_to(sd)
    except ValueError:
        raise FileNotFoundError("Некорректное имя файла настроек.") from None
    rel = _rel_parser_settings_env_name(safe_name)
    project_root = _project_root()
    env_path = current_app.extensions["env_path"]
    io_cfg = load_app_runtime(project_root, read_env_file(env_path))
    new_io = io_cfg.model_copy(update={"parser_settings_path": rel})
    save_app_runtime(project_root, new_io)
    apply_app_runtime_to_environ(new_io)
    current_app.extensions["app_runtime_config"] = new_io
    current_app.config["PARSER_SETTINGS_PATH"] = rel
    configure_settings_path(target)
    return target


def _build_non_overwriting_settings_path(filename: str) -> Path:
    safe_name = Path(filename).name
    base = Path(safe_name).stem
    suffix = Path(safe_name).suffix or ".json"
    target = _settings_dir() / safe_name
    if not target.exists():
        return target
    idx = 1
    while True:
        candidate = _settings_dir() / f"{base}_{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def _settings_page_context(io_cfg_dict: dict[str, Any], parser_text: str) -> dict:
    er_repo = EmergencyRepository(current_app.extensions["session_factory"])
    parser_path = _effective_parser_settings_path()
    choices, settings_files_warning = _settings_file_choices(parser_path.name)
    timezone_choices = [
        ("UTC", "UTC"),
        ("Europe/Kaliningrad", "UTC+2  Europe/Kaliningrad"),
        ("Europe/Moscow", "UTC+3  Europe/Moscow"),
        ("Europe/Samara", "UTC+4  Europe/Samara"),
        ("Asia/Yekaterinburg", "UTC+5  Asia/Yekaterinburg"),
        ("Asia/Omsk", "UTC+6  Asia/Omsk"),
        ("Asia/Krasnoyarsk", "UTC+7  Asia/Krasnoyarsk"),
        ("Asia/Irkutsk", "UTC+8  Asia/Irkutsk"),
        ("Asia/Yakutsk", "UTC+9  Asia/Yakutsk"),
        ("Asia/Vladivostok", "UTC+10 Asia/Vladivostok"),
        ("Asia/Magadan", "UTC+11 Asia/Magadan"),
        ("Asia/Kamchatka", "UTC+12 Asia/Kamchatka"),
    ]
    tz_select, tz_custom = _timezone_form_fields({"APP_TIMEZONE": io_cfg_dict.get("app_timezone", "")}, timezone_choices)
    io_display = dict(io_cfg_dict)
    io_display["_tz_select"] = tz_select
    fm_url = str(io_display.get("file_manager_url", "") or "").strip()
    window_raw = io_display.get("video_match_window_minutes", 20)
    try:
        window_minutes = int(window_raw)
    except Exception:
        window_minutes = 20
    return {
        "io_values": io_display,
        "app_timezone_custom_value": tz_custom,
        "parser_text": parser_text,
        "settings_files": choices,
        "active_settings_file": parser_path.name,
        "settings_files_warning": settings_files_warning,
        "emergency_conditions": er_repo.list_conditions(),
        "timezone_choices": timezone_choices,
        "file_manager_url": fm_url,
        "video_match_window_minutes": window_minutes,
    }


def _posted_io_dict_for_template() -> dict[str, Any]:
    f = request.form
    effective_tz = (f.get("app_timezone_custom") or "").strip() or (f.get("app_timezone_select") or "").strip()
    sf = (f.get("settings_file") or "settings.json").strip()
    return {
        "modbus_port": f.get("modbus_port", ""),
        "modbus_slave": f.get("modbus_slave", ""),
        "modbus_baudrate": f.get("modbus_baudrate", ""),
        "modbus_timeout": f.get("modbus_timeout", ""),
        "modbus_interval": f.get("modbus_interval", ""),
        "modbus_address_offset": f.get("modbus_address_offset", ""),
        "ram_batch_size": f.get("ram_batch_size", ""),
        "app_timezone": effective_tz,
        "parser_settings_path": _rel_parser_settings_env_name(sf),
        "disable_modbus_collector": f.get("disable_modbus_collector") == "1",
        "video_match_window_minutes": f.get("video_match_window_minutes", "20"),
        "file_manager_url": f.get("file_manager_url", ""),
    }


def _instructions_dir() -> Path:
    return _project_root() / "instructions"


_INSTRUCTION_FILES: dict[str, str] = {
    "emergency-rules": "emergency-rules.md",
    "settings-json": "settings-json.md",
}


@main_router.route("/", methods=["GET"])
def index():
    if current_user.is_authenticated:
        return redirect(url_for("main_blueprint.dashboard"))
    return redirect(url_for("auth_blueprint.login"))


@main_router.route("/dashboard", methods=["GET"])
@login_required
def dashboard():
    analog_opts = analog_labels_for(all_analog_keys())
    discrete_opts = discrete_labels_for(all_discrete_keys())
    return render_template(
        "dashboard/index.html",
        analog_options=analog_opts,
        discrete_options=discrete_opts,
    )


@main_router.route("/settings", methods=["GET"])
@login_required
def settings():
    io_cfg = current_app.extensions["app_runtime_config"]
    parser_path = _effective_parser_settings_path()
    with parser_path.open("r", encoding="utf-8") as f:
        parser_text = f.read()
    return render_template("settings/index.html", **_settings_page_context(io_cfg.model_dump(), parser_text))


@main_router.route("/settings/gpio", methods=["GET"])
@login_required
def settings_gpio():
    project_root = _project_root()
    path = gpio_inputs_path(project_root)
    ensure_gpio_inputs_file(path)
    try:
        text = path.read_text(encoding="utf-8")
    except OSError as exc:
        flash(f"Не удалось прочитать настройки GPIO: {exc}", "error")
        text = ""
    is_admin = _is_admin()
    return render_template("settings/gpio.html", gpio_text=text, gpio_path=path, is_admin=is_admin)


@main_router.route("/settings/gpio", methods=["POST"])
@login_required
def settings_gpio_save():
    if not _is_admin():
        flash("Недостаточно прав для этого действия.", "error")
        return redirect(url_for("main_blueprint.settings_gpio"))

    project_root = _project_root()
    path = gpio_inputs_path(project_root)
    ensure_gpio_inputs_file(path)
    text = request.form.get("gpio_json") or ""
    _cfg, err = validate_gpio_inputs_json(text)
    if err:
        flash(f"Настройки GPIO не сохранены: {err}", "error")
        return render_template("settings/gpio.html", gpio_text=text, gpio_path=path, is_admin=True), 400
    path.write_text(text.strip() + "\n", encoding="utf-8")
    collector = current_app.extensions.get("gpio_collector")
    try:
        if collector is not None:
            collector.restart(gpio_settings_path=path)
            flash("Настройки GPIO сохранены. GPIO-ридер перезапущен.", "success")
        else:
            flash("Настройки GPIO сохранены.", "success")
    except Exception as exc:
        flash(f"Настройки GPIO сохранены, но перезапуск GPIO-ридера не удался: {exc}", "error")
    return redirect(url_for("main_blueprint.settings_gpio"))


@main_router.route("/admin/event-logs", methods=["GET"])
@admin_required
def event_logs_page():
    repo = DataRepository(current_app.extensions["session_factory"])
    rows = repo.list_event_logs(limit=500)
    try:
        return render_template("settings/event_logs.html", rows=rows)
    except TemplateNotFound:
        return render_template_string(
            """
{% extends "base.html" %}
{% block title %}Логи событий · BlackBox{% endblock %}
{% block content %}
<section class="card">
    <h2>Логи событий системы</h2>
    <p class="hint">Показаны последние 500 записей (новые сверху).</p>
    <div class="table-wrap">
        <table>
            <thead>
            <tr><th>Время</th><th>Уровень</th><th>Код</th><th>Сообщение</th><th>Payload</th></tr>
            </thead>
            <tbody>
            {% for row in rows %}
                <tr>
                    <td>{{ format_in_configured_timezone(row.created_at, "%d.%m.%Y %H:%M:%S") }}</td>
                    <td>{{ row.level }}</td>
                    <td>{{ row.code }}</td>
                    <td>{{ row.message }}</td>
                    <td><code>{{ row.payload_json or '' }}</code></td>
                </tr>
            {% endfor %}
            </tbody>
        </table>
    </div>
    {% if not rows %}<p class="hint">Записей пока нет.</p>{% endif %}
</section>
{% endblock %}
""",
            rows=rows,
        )


@main_router.route("/settings/instructions/<slug>", methods=["GET"])
@admin_required
def settings_instruction(slug: str):
    filename = _INSTRUCTION_FILES.get(str(slug).strip().lower())
    if not filename:
        return Response("Инструкция не найдена.", status=404, mimetype="text/plain")
    path = _instructions_dir() / filename
    if not path.exists():
        return Response("Файл инструкции отсутствует.", status=404, mimetype="text/plain")
    text = path.read_text(encoding="utf-8")
    return Response(text, status=200, mimetype="text/plain")


@main_router.route("/settings", methods=["POST"])
@login_required
def settings_save():
    env_path = current_app.extensions["env_path"]
    env_disk = read_env_file(env_path)
    project_root = _project_root()
    parser_path = _effective_parser_settings_path()
    static_csv_dir = current_app.extensions["static_csv_dir"]
    collector = current_app.extensions["modbus_collector"]

    parser_text = request.form.get("parser_json") or ""
    action = (request.form.get("action") or "test").strip().lower()
    selected_settings_file = (request.form.get("settings_file") or parser_path.name).strip()
    is_admin = _is_admin()

    def _fail(parser_txt: str | None = None, *, status: int = 400):
        text = parser_txt if parser_txt is not None else parser_text
        io_d = _posted_io_dict_for_template()
        return render_template("settings/index.html", **_settings_page_context(io_d, text)), status

    if action == "upload":
        if not is_admin:
            flash("Недостаточно прав для этого действия.", "error")
            return _fail(status=403)
        uploaded = request.files.get("settings_file_upload")
        if uploaded is None or not uploaded.filename:
            flash("Выберите JSON-файл настроек.", "error")
            return _fail()
        filename = Path(uploaded.filename).name
        if not filename.lower().endswith(".json"):
            flash("Допустимы только файлы .json", "error")
            return _fail()
        try:
            raw = uploaded.read().decode("utf-8")
            payload = json.loads(raw)
        except Exception as exc:
            flash(f"Ошибка чтения файла импорта: {exc}", "error")
            return _fail()

        if not isinstance(payload, dict):
            flash("Файл настроек должен быть JSON-объектом.", "error")
            return _fail()

        parser_cfg_import, err = validate_parser_json(raw)
        if err:
            flash(f"Файл не загружен: {err}", "error")
            return _fail()
        _ = parser_cfg_import
        target = _build_non_overwriting_settings_path(filename)
        if target.name == APP_RUNTIME_FILENAME:
            flash(f"Имя «{APP_RUNTIME_FILENAME}» зарезервировано под настройки Modbus.", "error")
            return _fail()
        target.write_text(raw.strip() + "\n", encoding="utf-8")
        _set_active_parser_settings(target.name)
        reload_settings_cache()
        collector.restart()
        if target.name == filename:
            flash(f"Файл '{target.name}' загружен и выбран активным. Чтение перезапущено.", "success")
        else:
            flash(
                f"Файл '{filename}' уже существовал. Загружен как '{target.name}' и выбран активным. Чтение перезапущено.",
                "success",
            )
        return redirect(url_for("main_blueprint.settings"))

    if action == "switch_file":
        try:
            target_path = _settings_dir() / Path(selected_settings_file).name
            if not target_path.exists():
                raise FileNotFoundError(f"Файл настроек не найден: {Path(selected_settings_file).name}")
            if target_path.name == APP_RUNTIME_FILENAME:
                raise FileNotFoundError(f"Нельзя выбрать служебный файл {APP_RUNTIME_FILENAME}.")
        except Exception as exc:
            flash(str(exc), "error")
            return _fail()
        parser_text = target_path.read_text(encoding="utf-8")
        parser_cfg_selected, err = validate_parser_json(parser_text)
        if err:
            flash(f"Выбранный файл невалиден: {err}", "error")
            return _fail()
        _ = parser_cfg_selected
        _set_active_parser_settings(target_path.name)
        reload_settings_cache()
        collector.restart()
        flash("Активный файл настроек изменен. Чтение перезапущено.", "success")
        return redirect(url_for("main_blueprint.settings"))

    if not is_admin:
        if action != "save":
            flash("Недостаточно прав для этого действия.", "error")
            return _fail(status=403)
        effective_tz = (request.form.get("app_timezone_custom") or "").strip() or (
            request.form.get("app_timezone_select") or ""
        ).strip()
        target_path = _settings_dir() / Path(selected_settings_file).name
        if not target_path.exists() or target_path.name == APP_RUNTIME_FILENAME:
            flash("Некорректный файл настроек парсера.", "error")
            return _fail()
        parser_cfg_selected, err = validate_parser_json(target_path.read_text(encoding="utf-8"))
        _ = parser_cfg_selected
        if err:
            flash(f"Выбранный файл невалиден: {err}", "error")
            return _fail()

        project_root = _project_root()
        current_cfg = current_app.extensions["app_runtime_config"]
        new_cfg = current_cfg.model_copy(
            update={
                "app_timezone": effective_tz or current_cfg.app_timezone,
                "parser_settings_path": _rel_parser_settings_env_name(Path(selected_settings_file).name),
            }
        )
        save_app_runtime(project_root, new_cfg)
        apply_app_runtime_to_environ(new_cfg)
        current_app.extensions["app_runtime_config"] = new_cfg
        current_app.config["PARSER_SETTINGS_PATH"] = new_cfg.parser_settings_path
        configure_settings_path(target_path.resolve())
        reload_settings_cache()
        collector.restart()
        flash("Изменены часовой пояс и активный файл парсера.", "success")
        return redirect(url_for("main_blueprint.settings"))

    try:
        effective_tz = (request.form.get("app_timezone_custom") or "").strip() or (
            request.form.get("app_timezone_select") or ""
        ).strip()
        parser_rel = _rel_parser_settings_env_name(selected_settings_file)
        io_cfg = io_form_to_runtime(
            modbus_port=request.form.get("modbus_port") or "",
            modbus_slave=(request.form.get("modbus_slave") or "1").strip(),
            modbus_baudrate=(request.form.get("modbus_baudrate") or "9600").strip(),
            modbus_timeout=(request.form.get("modbus_timeout") or "0.35").strip(),
            modbus_interval=(request.form.get("modbus_interval") or "0.12").strip(),
            modbus_address_offset=(request.form.get("modbus_address_offset") or "1").strip(),
            ram_batch_size=(request.form.get("ram_batch_size") or "60").strip(),
            app_timezone=effective_tz or "Europe/Moscow",
            parser_settings_path=parser_rel,
            disable_modbus_collector=request.form.get("disable_modbus_collector") == "1",
            video_match_window_minutes=(request.form.get("video_match_window_minutes") or "20").strip(),
            file_manager_url=(request.form.get("file_manager_url") or "").strip(),
        )
    except Exception as exc:
        flash(f"Некорректные настройки чтения данных: {exc}", "error")
        return _fail()

    parser_cfg, err = validate_parser_json(parser_text)
    if err:
        flash(err, "error")
        return _fail()

    db_path_raw = env_disk.get("BLACKBOX_DB_PATH", ROOT_ENV_DEFAULTS["BLACKBOX_DB_PATH"])
    try:
        runtime = build_runtime_config(io_cfg, db_path=db_path_raw, static_csv_dir=static_csv_dir)
    except Exception as exc:
        flash(f"Ошибка сборки конфигурации: {exc}", "error")
        return _fail()

    collector.stop()
    ok, msg = test_modbus_settings(runtime, parser_cfg)
    if not ok:
        collector.start()
        flash(msg, "error")
        return _fail()

    if action == "test":
        collector.start()
        flash(msg, "success")
        io_d = _posted_io_dict_for_template()
        return render_template("settings/index.html", **_settings_page_context(io_d, parser_text))

    if action != "save":
        collector.start()
        flash("Неизвестное действие.", "error")
        return _fail()

    sd = _settings_dir().resolve()
    try:
        target_write = (sd / Path(selected_settings_file).name).resolve()
        target_write.relative_to(sd)
        if target_write.name == APP_RUNTIME_FILENAME:
            raise FileNotFoundError(f"Нельзя записывать парсер в {APP_RUNTIME_FILENAME}.")
    except Exception as exc:
        collector.start()
        flash(str(exc), "error")
        return _fail()

    try:
        with target_write.open("w", encoding="utf-8") as f:
            f.write(parser_text.strip() + "\n")
    except OSError as exc:
        collector.start()
        flash(f"Не удалось записать JSON парсера: {exc}", "error")
        return _fail()

    save_app_runtime(project_root, io_cfg)
    apply_app_runtime_to_environ(io_cfg)
    current_app.extensions["app_runtime_config"] = io_cfg
    current_app.config["PARSER_SETTINGS_PATH"] = io_cfg.parser_settings_path
    configure_settings_path(target_write)
    reload_settings_cache()

    if io_cfg.disable_modbus_collector:
        collector.stop()
    else:
        collector.restart(new_config=runtime)

    flash("Настройки чтения сохранены в app_runtime.json; процесс чтения перезапущен.", "success")
    return redirect(url_for("main_blueprint.settings"))


@main_router.route("/alarms", methods=["GET"])
@login_required
def alarms_page():
    er_repo = EmergencyRepository(current_app.extensions["session_factory"])
    emergency_rows = er_repo.list_recent_emergencies(limit=200)
    analog_opts = analog_labels_for(all_analog_keys())
    discrete_opts = discrete_labels_for(all_discrete_keys())
    return render_template(
        "alarms/index.html",
        emergency_rows=emergency_rows,
        analog_options=analog_opts,
        discrete_options=discrete_opts,
    )


def _alarm_export_rows(
    *,
    table: str,
    rows_db: list[Any],
    analog_keys: list[str],
    discrete_keys: list[str],
) -> tuple[list[str], list[list[Any]]]:
    if table == "alarms":
        headers = ["Дата", "Время", "Название", "Состояние"]
        body = [
            [item.created_at.strftime("%d/%m/%Y"), item.created_at.strftime("%H:%M:%S"), item.name, getattr(item, "state", "active")]
            for item in rows_db
        ]
        return headers, body

    if table == "analog":
        headers = ["Дата", "Время", *analog_keys]
        body: list[list[Any]] = []
        for item in rows_db:
            processed = decode_to_processed(item.date)
            analog, _ = analog_discrete_for_csv(processed)
            body.append(
                [
                    item.created_at.strftime("%d/%m/%Y"),
                    item.created_at.strftime("%H:%M:%S"),
                    *[analog.get(k, "") for k in analog_keys],
                ]
            )
        return headers, body

    headers = ["Дата", "Время", *discrete_keys]
    body = []
    for item in rows_db:
        processed = decode_to_processed(item.date)
        _, discrete = analog_discrete_for_csv(processed)
        body.append(
            [
                item.created_at.strftime("%d/%m/%Y"),
                item.created_at.strftime("%H:%M:%S"),
                *[1 if bool(discrete.get(k, False)) else 0 for k in discrete_keys],
            ]
        )
    return headers, body


@main_router.route("/alarms/export", methods=["POST"])
@login_required
def alarms_export():
    try:
        event_id = int((request.form.get("event_id") or "").strip())
    except Exception:
        return render_template("data/_export_result.html", error="Некорректный идентификатор события.", download_url=None), 400
    er_repo = EmergencyRepository(current_app.extensions["session_factory"])
    event = er_repo.get_emergency_event(event_id)
    if event is None:
        return render_template("data/_export_result.html", error="Событие не найдено.", download_url=None), 404

    tables: list[str] = []
    if request.form.get("table_analog") == "1":
        tables.append("analog")
    if request.form.get("table_discrete") == "1":
        tables.append("discrete")
    if request.form.get("table_alarms") == "1":
        tables.append("alarms")
    if not tables:
        return render_template("data/_export_result.html", error="Выберите хотя бы одну таблицу для экспорта.", download_url=None)

    export_format = (request.form.get("export_format") or "csv_zip").strip().lower()
    if export_format not in ("csv_zip", "excel"):
        return render_template("data/_export_result.html", error="Некорректный формат экспорта.", download_url=None)

    sort_desc = (request.form.get("sort") or "desc").lower() != "asc"
    analog_keys = filter_valid_analog(request.form.getlist("analog_col_export") or None)
    discrete_keys = filter_valid_discrete(request.form.getlist("discrete_col_export") or None)
    if "analog" in tables and not analog_keys:
        return render_template("data/_export_result.html", error="Для таблицы «Аналоги» выберите хотя бы одно поле.", download_url=None)
    if "discrete" in tables and not discrete_keys:
        return render_template("data/_export_result.html", error="Для таблицы «Дискреты» выберите хотя бы одно поле.", download_url=None)

    window_from = event.datetime - timedelta(minutes=10)
    window_to = (event.ended_at or event.datetime) + timedelta(minutes=10)

    repo = DataRepository(current_app.extensions["session_factory"])
    collector = current_app.extensions["modbus_collector"]
    if "alarms" in tables and not collector._alarms_enabled:
        return render_template(
            "data/_export_result.html",
            error="Таблица аварий недоступна (нет таблицы в БД). Выполните миграции.",
            download_url=None,
        )

    static_csv_dir: Path = current_app.extensions["static_csv_dir"]
    static_csv_dir.mkdir(parents=True, exist_ok=True)
    stamp = event.datetime.strftime("%Y%m%d_%H%M%S")
    dataset: dict[str, tuple[list[str], list[list[Any]]]] = {}
    video_meta_rows: list[list[Any]] = []
    video_files_to_add: list[tuple[str, str]] = []  # (absolute_path, arcname)

    def _video_allowed_roots() -> list[Path]:
        env = read_env_file(current_app.extensions["env_path"])
        raw = (env.get("VIDEO_STORAGE_DIR") or "").strip()
        if not raw:
            return []
        root = Path(raw).expanduser().resolve()
        roots = [root]
        if re.fullmatch(r"cam\d+", root.name.lower()) and root.parent.is_dir():
            roots.append(root.parent.resolve())
        uniq: list[Path] = []
        for p in roots:
            if p not in uniq:
                uniq.append(p)
        return uniq

    def _normalize_video_fs_path(raw: str) -> Path:
        value = str(raw or "").strip()
        lower = value.lower()
        if lower.startswith("file="):
            value = value.split("=", 1)[1].strip()
        elif lower.startswith("file://"):
            value = value[7:].strip()
        return Path(value).expanduser()

    def _is_under_any_root(path: Path, roots: list[Path]) -> bool:
        if not roots:
            return True
        for root in roots:
            try:
                path.relative_to(root)
                return True
            except ValueError:
                continue
        return False

    for table in tables:
        if table == "analog":
            rows_db = repo.list_analogs(created_from=window_from, created_to=window_to, sort_desc=sort_desc, offset=0, limit=None)
        elif table == "discrete":
            rows_db = repo.list_discretes(created_from=window_from, created_to=window_to, sort_desc=sort_desc, offset=0, limit=None)
        else:
            rows_db = repo.list_alarms(created_from=window_from, created_to=window_to, sort_desc=sort_desc, offset=0, limit=None)
        dataset[table] = _alarm_export_rows(
            table=table,
            rows_db=rows_db,
            analog_keys=analog_keys,
            discrete_keys=discrete_keys,
        )

    # Если экспортируются аварии — добавляем в пакет связанные видео по alarm_id.
    if "alarms" in tables:
        session_factory = current_app.extensions["session_factory"]
        allowed_video_roots = _video_allowed_roots()
        with session_factory() as session:
            video_q = (
                session.query(Video)
                .join(Alarms, Video.alarm_id == Alarms.id)
                .filter(Video.alarm_id.is_not(None))
                .filter(Alarms.created_at >= window_from, Alarms.created_at <= window_to)
                .order_by(Video.created_at.desc() if sort_desc else Video.created_at.asc())
            )
            for v in video_q.yield_per(1000):
                video_meta_rows.append(
                    [
                        v.id,
                        v.created_at.strftime("%d/%m/%Y"),
                        v.created_at.strftime("%H:%M:%S"),
                        v.captured_at.strftime("%d/%m/%Y %H:%M:%S"),
                        v.file_name,
                        v.file_path,
                        v.alarm_id if v.alarm_id is not None else "",
                    ]
                )
                file_path = _normalize_video_fs_path(v.file_path)
                try:
                    resolved = file_path.resolve()
                except OSError:
                    continue
                if not resolved.exists() or not resolved.is_file():
                    continue
                if not _is_under_any_root(resolved, allowed_video_roots):
                    continue
                arcname = f"videos/{v.id}_{v.file_name}"
                video_files_to_add.append((str(resolved), arcname))

    if export_format == "csv_zip":
        out_path = static_csv_dir / f"alarm_event_export_{event_id}_{stamp}.zip"
        with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for table in tables:
                headers, body = dataset[table]
                sio = io.StringIO()
                writer = csv.writer(sio, delimiter=";")
                writer.writerow(headers)
                writer.writerows(body)
                zf.writestr(f"{table}.csv", sio.getvalue().encode("utf-8-sig"))

            if video_meta_rows:
                sio = io.StringIO()
                writer = csv.writer(sio, delimiter=";")
                writer.writerow(["ID", "Дата", "Время", "Время_захвата", "Имя_файла", "Путь_к_файлу", "ID_аварии"])
                writer.writerows(video_meta_rows)
                zf.writestr("videos.csv", sio.getvalue().encode("utf-8-sig"))
                for abs_path, arcname in video_files_to_add:
                    zf.write(abs_path, arcname=arcname)
        rel = out_path.relative_to(current_app.static_folder)
        return render_template(
            "data/_export_result.html",
            error=None,
            download_url=url_for("static", filename=str(rel).replace("\\", "/")),
        )

    try:
        from openpyxl import Workbook
    except Exception:
        return render_template(
            "data/_export_result.html",
            error="Для Excel-экспорта требуется пакет openpyxl. Установите зависимость и повторите.",
            download_url=None,
        )
    wb = Workbook()
    wb.remove(wb.active)
    sheet_title = {"analog": "Аналоги", "discrete": "Дискреты", "alarms": "Аварии"}
    for table in tables:
        ws = wb.create_sheet(sheet_title.get(table, table))
        headers, body = dataset[table]
        ws.append(headers)
        for row in body:
            ws.append(row)

    if video_meta_rows:
        ws = wb.create_sheet("Видео")
        ws.append(["ID", "Дата", "Время", "Время_захвата", "Имя_файла", "Путь_к_файлу", "ID_аварии"])
        for row in video_meta_rows:
            ws.append(row)
    out_path = static_csv_dir / f"alarm_event_export_{event_id}_{stamp}.xlsx"
    wb.save(out_path)
    rel = out_path.relative_to(current_app.static_folder)
    return render_template("data/_export_result.html", error=None, download_url=url_for("static", filename=str(rel).replace("\\", "/")))


@main_router.route("/settings/emergency-rules", methods=["POST"])
@admin_required
def emergency_rule_add():
    name = (request.form.get("rule_name") or "").strip()
    condition = (request.form.get("rule_condition") or "").strip()
    if not name:
        flash("Укажите название правила.", "error")
        return redirect(url_for("main_blueprint.settings"))
    parser_path = _effective_parser_settings_path()
    ok, err = validate_emergency_rule_expression(condition, settings_path=parser_path)
    if not ok:
        flash(err or "Некорректное правило.", "error")
        return redirect(url_for("main_blueprint.settings"))
    er_repo = EmergencyRepository(current_app.extensions["session_factory"])
    er_repo.create_condition(name=name, condition=condition)
    current_app.extensions["modbus_collector"].restart()
    flash("Правило аварии сохранено.", "success")
    return redirect(url_for("main_blueprint.settings"))


@main_router.route("/settings/emergency-rules/<int:rule_id>/edit", methods=["POST"])
@admin_required
def emergency_rule_edit(rule_id: int):
    name = (request.form.get("rule_name") or "").strip()
    condition = (request.form.get("rule_condition") or "").strip()
    if not name:
        flash("Укажите название правила.", "error")
        return redirect(url_for("main_blueprint.settings"))
    parser_path = _effective_parser_settings_path()
    ok, err = validate_emergency_rule_expression(condition, settings_path=parser_path)
    if not ok:
        flash(err or "Некорректное правило.", "error")
        return redirect(url_for("main_blueprint.settings"))
    er_repo = EmergencyRepository(current_app.extensions["session_factory"])
    if er_repo.update_condition(condition_id=rule_id, name=name, condition=condition):
        current_app.extensions["modbus_collector"].restart()
        flash("Правило обновлено.", "success")
    else:
        flash("Правило не найдено.", "error")
    return redirect(url_for("main_blueprint.settings"))


@main_router.route("/settings/emergency-rules/<int:rule_id>/delete", methods=["POST"])
@admin_required
def emergency_rule_delete(rule_id: int):
    er_repo = EmergencyRepository(current_app.extensions["session_factory"])
    if er_repo.soft_delete_condition(rule_id):
        current_app.extensions["modbus_collector"].restart()
        flash("Правило помечено удалённым.", "success")
    else:
        flash("Правило не найдено.", "error")
    return redirect(url_for("main_blueprint.settings"))


def _build_live_dashboard_context(
    analog_columns: list[str],
    discrete_columns: list[str],
) -> dict:
    session_factory = current_app.extensions["session_factory"]
    repo = DataRepository(session_factory)

    latest_rows = repo.list_analogs(limit=1)
    latest_row = latest_rows[0] if latest_rows else None
    latest_processed: dict[str, Any] = {}
    latest_time = None
    if latest_row is not None:
        latest_processed = decode_to_processed(latest_row.date)
        latest_time = format_in_configured_timezone(latest_row.created_at, DATETIME_UI_FORMAT)

    analog_items: list[dict] = []
    analog_time = latest_time
    analog_label_map = dict(analog_labels_for(analog_columns))
    if latest_row is not None:
        analog_map, _ = analog_discrete_for_csv(latest_processed)
        analog_items = [{"name": analog_label_map.get(k, k), "value": analog_map.get(k, "")} for k in analog_columns]

    discrete_items: list[dict] = []
    discrete_time = latest_time
    discrete_label_map = dict(discrete_labels_for(discrete_columns))
    if latest_row is not None:
        _, discrete_map = analog_discrete_for_csv(latest_processed)
        discrete_items = [
            {"name": discrete_label_map.get(k, k), "is_on": bool(discrete_map.get(k, False))}
            for k in discrete_columns
        ]

    active_alarms_raw = latest_processed.get("active_alarms", [])
    active_alarms = [str(v) for v in active_alarms_raw] if isinstance(active_alarms_raw, list) else []
    alarm_rows = [{"time": latest_time, "name": name} for name in active_alarms]

    system_monitor = _collect_system_monitor()

    return {
        "server_time": format_in_configured_timezone(datetime.now(), DATETIME_UI_FORMAT),
        "analog_time": analog_time,
        "analog_items": analog_items,
        "discrete_time": discrete_time,
        "discrete_items": discrete_items,
        "alarm_rows": alarm_rows,
        "alarm_time": latest_time,
        "system_monitor": system_monitor,
    }


def render_live_dashboard_html(analog_columns: list[str], discrete_columns: list[str]) -> str:
    ctx = _build_live_dashboard_context(analog_columns, discrete_columns)
    return render_template("dashboard/_live_panels.html", **ctx)


def _decode_off_fields(raw: str | None) -> set[str]:
    if raw is None:
        return set()
    text = str(raw).strip()
    if not text:
        return set()
    padded = text + ("=" * ((4 - len(text) % 4) % 4))
    try:
        decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        payload = json.loads(decoded)
        if not isinstance(payload, list):
            return set()
        return {str(v) for v in payload}
    except (ValueError, UnicodeDecodeError, binascii.Error):
        return set()


def _collect_system_monitor() -> dict[str, Any]:
    stats: dict[str, Any] = {
        "disk": {"used_gb": None, "total_gb": None, "free_gb": None, "percent": None},
        "cpu": {"percent": None, "cores_logical": None, "cores_physical": None, "fan_rpm": None},
        "memory": {"used_gb": None, "total_gb": None, "percent": None},
        "process": {"pid": os.getpid(), "uptime_sec": None},
        "disks": [],
    }
    try:
        du = shutil.disk_usage(_project_root())
        stats["disk"] = {
            "used_gb": round((du.total - du.free) / (1024**3), 2),
            "total_gb": round(du.total / (1024**3), 2),
            "free_gb": round(du.free / (1024**3), 2),
            "percent": round(((du.total - du.free) / du.total) * 100.0, 1) if du.total else 0.0,
        }
    except Exception:
        pass

    try:
        import psutil  # type: ignore

        vm = psutil.virtual_memory()
        proc = psutil.Process(os.getpid())
        stats["cpu"] = {
            "percent": round(psutil.cpu_percent(interval=None), 1),
            "cores_logical": int(psutil.cpu_count(logical=True) or 0),
            "cores_physical": int(psutil.cpu_count(logical=False) or 0),
        }
        try:
            fans = psutil.sensors_fans()
            selected_rpm = None
            for source_name, entries in fans.items():
                for entry in entries:
                    label = (getattr(entry, "label", "") or "").lower()
                    candidate = getattr(entry, "current", None)
                    if candidate is None:
                        continue
                    if "cpu" in label or "proc" in label or "cpu" in source_name.lower():
                        selected_rpm = int(candidate)
                        break
                    if selected_rpm is None:
                        selected_rpm = int(candidate)
                if selected_rpm is not None:
                    break
            stats["cpu"]["fan_rpm"] = selected_rpm
        except Exception:
            pass
        stats["memory"] = {
            "used_gb": round((vm.total - vm.available) / (1024**3), 2),
            "total_gb": round(vm.total / (1024**3), 2),
            "percent": round(float(vm.percent), 1),
        }
        stats["process"]["uptime_sec"] = int(max(0.0, (time.time() - proc.create_time())))
        disks: list[dict[str, Any]] = []
        seen_mounts: set[str] = set()
        for part in psutil.disk_partitions(all=False):
            mount = str(getattr(part, "mountpoint", "") or "")
            device = str(getattr(part, "device", "") or "")
            fstype = str(getattr(part, "fstype", "") or "")
            if not mount or mount in seen_mounts:
                continue
            seen_mounts.add(mount)
            try:
                usage = psutil.disk_usage(mount)
            except Exception:
                continue
            disks.append(
                {
                    "mount": mount,
                    "device": device,
                    "fstype": fstype,
                    "used_gb": round((usage.total - usage.free) / (1024**3), 2),
                    "total_gb": round(usage.total / (1024**3), 2),
                    "free_gb": round(usage.free / (1024**3), 2),
                    "percent": round(float(usage.percent), 1),
                }
            )
        stats["disks"] = disks
    except Exception:
        pass

    return stats


@main_router.route("/dashboard/live", methods=["GET"])
@login_required
def dashboard_live():
    analog_off = _decode_off_fields(request.args.get("analog_off_b64"))
    discrete_off = _decode_off_fields(request.args.get("discrete_off_b64"))
    analog_columns = [k for k in filter_valid_analog(None) if k not in analog_off]
    discrete_columns = [k for k in filter_valid_discrete(None) if k not in discrete_off]
    if not analog_columns:
        analog_columns = filter_valid_analog(None)
    if not discrete_columns:
        discrete_columns = filter_valid_discrete(None)
    return render_live_dashboard_html(analog_columns, discrete_columns)
